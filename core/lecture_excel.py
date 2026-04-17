"""
core/lecture_excel.py
Lecture du fichier Excel de saisie BAEL91_Saisie_v13.xlsx
→ construction d'un objet Projet complet

Structure du fichier :
  Feuille Materiaux  : col B = symbole, col C = valeur (lignes 3+)
  Feuille Noeuds     : en-têtes ligne 3, données ligne 4+
  Feuille Barres     : en-têtes ligne 3, données ligne 5+
                       (ligne 4 = sous-titre "POTEAUX", ignorée)
  Feuille Dalles     : en-têtes ligne 3, données ligne 4+
  Feuille Fondations : en-têtes ligne 3, données ligne 4+
"""
from typing import List, Tuple
from .declarations import Projet, Materiaux, Noeud, Barre, Dalle, Semelle


def lire_excel(chemin_ou_buffer) -> Tuple[Projet, List[str]]:
    """
    Lit un fichier Excel de saisie et retourne (Projet, liste_erreurs).
    Si liste_erreurs est vide, le projet est valide.
    """
    try:
        import openpyxl
    except ImportError:
        return Projet(), ["openpyxl non installé — pip install openpyxl"]

    erreurs = []

    try:
        wb = openpyxl.load_workbook(chemin_ou_buffer, data_only=True)
    except Exception as e:
        return Projet(), [f"Impossible d'ouvrir le fichier : {e}"]

    feuilles = wb.sheetnames

    # ── 1. Matériaux ──────────────────────────────────────────────────────────
    mat = Materiaux()
    if "Materiaux" in feuilles:
        ws = wb["Materiaux"]
        sym_map = {
            "fc28":     "fc28",
            "gammab":   "gammab",
            "rba":      "rhoba",
            "fe":       "fe",
            "gammas":   "gammas",
            "c_poutre": "c_poutre",
            "c_dalle":  "c_dalle",
            "c_poteau": "c_poteau",
            "c_fond":   "c_fond",
            "Df":       "Df",
            "q_adm":    "q_adm",
        }
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
            sym = str(row[1]).strip() if row[1] else ""
            val = row[2]
            if sym in sym_map and val is not None:
                try:
                    # Ignorer les formules Excel (str commençant par =)
                    if isinstance(val, str) and val.startswith("="):
                        continue
                    setattr(mat, sym_map[sym], float(val))
                except (ValueError, TypeError):
                    pass
        # Classe d'exposition (chercher dans col B)
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
            if row[1] and "classe" in str(row[1]).lower():
                if row[2]:
                    mat.classe_exposition = str(row[2]).strip()

        # ── Vérification et correction des valeurs matériaux ──────────────────
        # Détection des formules Excel non évaluées par openpyxl
        # (fbu, fsu, sigmaBc, ftj sont des propriétés calculées depuis fc28/fe)
        # Si les valeurs lues sont hors plage physique → recalculer depuis fc28/fe
        params_corriges = []

        def _hors_plage(val, lo, hi):
            try:
                return not (lo <= float(val) <= hi)
            except (TypeError, ValueError):
                return True

        # fbu = 0.85 × fc28 / gammab → plage [5, 35] MPa
        if _hors_plage(mat.fbu, 5.0, 35.0):
            params_corriges.append(
                f"fbu (lu={mat.fbu:.4f} MPa, recalculé={mat.fbu:.2f}→{mat.fc28*0.85/mat.gammab:.2f} MPa)"
            )
        # fsu = fe / gammas → plage [150, 500] MPa
        if _hors_plage(mat.fsu, 150.0, 500.0):
            params_corriges.append(
                f"fsu (lu≈{mat.fsu:.2f} MPa, recalculé={mat.fe/mat.gammas:.2f} MPa)"
            )
        # sigmaBc = 0.6 × fc28 → plage [5, 30] MPa
        if _hors_plage(mat.sigmaBc, 5.0, 30.0):
            params_corriges.append(
                f"sigmaBc (lu={mat.sigmaBc:.2f} MPa, recalculé={0.6*mat.fc28:.2f} MPa)"
            )
        # ftj = 0.6 + 0.06 × fc28 → plage [0.5, 5.0] MPa
        if _hors_plage(mat.ftj, 0.5, 5.0):
            params_corriges.append(
                f"ftj (lu={mat.ftj:.2f} MPa, recalculé={0.6+0.06*mat.fc28:.2f} MPa)"
            )
        # fc28 plausible → plage [10, 60] MPa
        if _hors_plage(mat.fc28, 10.0, 60.0):
            erreurs.append(
                f"fc28={mat.fc28} MPa hors plage [10-60] — vérifier la saisie"
            )
        # fe plausible → plage [200, 600] MPa
        if _hors_plage(mat.fe, 200.0, 600.0):
            erreurs.append(
                f"fe={mat.fe} MPa hors plage [200-600] — vérifier la saisie"
            )

        if params_corriges:
            msg = (
                "⚠ Valeurs matériaux incohérentes détectées dans Excel "
                "(formules non évaluées par openpyxl). "
                "Valeurs recalculées automatiquement depuis fc28 et fe : "
                + " | ".join(params_corriges)
                + ". Le calcul utilisera les valeurs théoriques BAEL."
            )
            erreurs.append(msg)
            # Note : fbu, fsu, sigmaBc, ftj sont des @property dans Materiaux
            # → elles sont TOUJOURS recalculées depuis fc28/fe automatiquement
            # → aucune correction de code nécessaire, l'avertissement suffit

    else:
        erreurs.append("Feuille 'Materiaux' introuvable — valeurs par défaut utilisées")

    # ── 2. Nœuds ──────────────────────────────────────────────────────────────
    noeuds = []
    if "Noeuds" in feuilles:
        ws = wb["Noeuds"]
        ids_noeuds = set()
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
            if not row[0]:
                continue
            try:
                nid = int(row[0])
                x   = float(row[1]) if row[1] is not None else 0.0
                y   = float(row[2]) if row[2] is not None else 0.0
                z   = float(row[3]) if row[3] is not None else 0.0
            except (ValueError, TypeError):
                continue
            if nid in ids_noeuds:
                erreurs.append(f"Nœud ID={nid} dupliqué — ignoré")
                continue
            ids_noeuds.add(nid)
            noeuds.append(Noeud(id=nid, x=x, y=y, z=z))
    else:
        erreurs.append("Feuille 'Noeuds' introuvable")

    if not noeuds:
        erreurs.append("Aucun nœud trouvé dans la feuille Noeuds")

    # ── 3. Barres ─────────────────────────────────────────────────────────────
    barres = []
    if "Barres" in feuilles:
        ws = wb["Barres"]
        ids_barres = set()
        ids_noeuds_set = {n.id for n in noeuds}
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
            # Ignorer les lignes de sous-titre (valeur texte en col A)
            if not row[0]:
                continue
            try:
                bid = int(float(row[0]))
            except (ValueError, TypeError):
                continue
            if bid in ids_barres:
                erreurs.append(f"Barre ID={bid} dupliquée — ignorée")
                continue
            try:
                nom   = str(row[1]).strip() if row[1] else f"B{bid}"
                ni    = int(float(row[2])) if row[2] is not None else 0
                nj    = int(float(row[3])) if row[3] is not None else 0
                b     = float(row[4]) if row[4] is not None else 0.25
                h     = float(row[5]) if row[5] is not None else 0.40
                G_add = float(row[6]) if row[6] is not None else 0.0
                Q_add = float(row[7]) if row[7] is not None else 0.0
            except (ValueError, TypeError) as e:
                erreurs.append(f"Barre ligne {bid} : données invalides ({e})")
                continue

            # Vérifications
            if ni not in ids_noeuds_set:
                erreurs.append(f"Barre {bid} : nœud Ni={ni} inexistant")
            if nj not in ids_noeuds_set:
                erreurs.append(f"Barre {bid} : nœud Nj={nj} inexistant")
            if ni == nj:
                erreurs.append(f"Barre {bid} : Ni=Nj={ni} (barre de longueur nulle)")
                continue
            if b <= 0 or h <= 0:
                erreurs.append(f"Barre {bid} : section b={b} h={h} invalide")
                continue

            ids_barres.add(bid)
            barres.append(Barre(
                id=bid, nom=nom, ni=ni, nj=nj,
                b=b, h=h, G_add=G_add, Q_add=Q_add
            ))
    else:
        erreurs.append("Feuille 'Barres' introuvable")

    # ── 4. Dalles ─────────────────────────────────────────────────────────────
    dalles = []
    if "Dalles" in feuilles:
        ws = wb["Dalles"]
        ids_dalles = set()
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
            if not row[0]:
                continue
            try:
                did = int(float(row[0]))
            except (ValueError, TypeError):
                continue
            if did in ids_dalles:
                continue

            # Nœuds : colonnes 1..5 (N1..N5)
            noeuds_dalle = []
            for col in range(1, 6):
                v = row[col]
                if v is not None:
                    try:
                        nid = int(float(v))
                        if nid > 0:
                            noeuds_dalle.append(nid)
                    except (ValueError, TypeError):
                        pass

            if len(noeuds_dalle) < 3:
                erreurs.append(f"Dalle {did} : moins de 3 nœuds")
                continue

            try:
                sens_raw  = str(row[6]).strip() if row[6] else "Sens X"
                G         = float(row[7]) if row[7] is not None else 0.0
                Q         = float(row[8]) if row[8] is not None else 0.0
                type_raw  = str(row[9]).strip() if row[9] else "Hourdis"
                e_raw     = row[10]
            except (ValueError, TypeError) as e:
                erreurs.append(f"Dalle {did} : données invalides ({e})")
                continue

            # Normaliser le sens
            sens = _normaliser_sens(sens_raw)

            # Normaliser le type
            if "pleine" in type_raw.lower():
                type_dalle = "Pleine"
                try:
                    e_dalle = float(e_raw) if e_raw else 0.15
                except (ValueError, TypeError):
                    e_dalle = 0.15
                    erreurs.append(f"Dalle {did} (pleine) : épaisseur invalide, 0.15m utilisée")
            else:
                type_dalle = "Hourdis"
                e_dalle    = 0.0

            # Avertissement G poids propre
            if G < 2.0:
                erreurs.append(
                    f"Dalle {did} : G={G:.1f} kN/m² semble faible. "
                    "Rappel : G doit inclure le poids propre de la dalle."
                )

            ids_dalles.add(did)
            dalles.append(Dalle(
                id=did, noeuds=noeuds_dalle,
                G=G, Q=Q,
                sens_lx=sens,
                type_dalle=type_dalle,
                e_dalle=e_dalle,
            ))
    else:
        erreurs.append("Feuille 'Dalles' introuvable")

    # ── 5. Fondations ─────────────────────────────────────────────────────────
    semelles = []
    if "Fondations" in feuilles:
        ws = wb["Fondations"]
        ids_sem = set()
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
            if not row[0]:
                continue
            try:
                id_pot = int(float(row[0]))
            except (ValueError, TypeError):
                continue
            if id_pot in ids_sem:
                continue

            def _f(v, default=0.0):
                try:
                    return float(v) if v is not None else default
                except (ValueError, TypeError):
                    return default

            def _i(v, default=0):
                try:
                    return int(float(v)) if v is not None else default
                except (ValueError, TypeError):
                    return default

            ex         = _f(row[1])
            ey         = _f(row[2])
            q_adm_loc  = _f(row[3])
            long_X_v   = _i(row[4])
            b_lX       = _f(row[5], 0.25)
            h_lX       = _f(row[6], 0.40)
            long_Y_v   = _i(row[7])
            b_lY       = _f(row[8], 0.25)
            h_lY       = _f(row[9], 0.40)

            # Avertissement longrine manquante
            if ex != 0 and long_X_v == 0:
                erreurs.append(
                    f"Semelle C{id_pot} : ex={ex:.3f}m mais Long_X_vers=0. "
                    "La longrine X ne sera pas calculée."
                )
            if ey != 0 and long_Y_v == 0:
                erreurs.append(
                    f"Semelle C{id_pot} : ey={ey:.3f}m mais Long_Y_vers=0. "
                    "La longrine Y ne sera pas calculée."
                )

            ids_sem.add(id_pot)
            semelles.append(Semelle(
                id_poteau=id_pot,
                ex=ex, ey=ey,
                q_adm_loc=q_adm_loc,
                long_X_vers=long_X_v,
                b_long_X=b_lX, h_long_X=h_lX,
                long_Y_vers=long_Y_v,
                b_long_Y=b_lY, h_long_Y=h_lY,
            ))

    # ── Construction du projet ─────────────────────────────────────────────────
    # Nom du projet depuis le nom du fichier si disponible
    nom_projet = "Projet importé"
    try:
        if hasattr(chemin_ou_buffer, 'name'):
            import os
            nom_projet = os.path.splitext(
                os.path.basename(chemin_ou_buffer.name)
            )[0]
        elif isinstance(chemin_ou_buffer, str):
            import os
            nom_projet = os.path.splitext(
                os.path.basename(chemin_ou_buffer)
            )[0]
    except Exception:
        pass

    projet = Projet(
        nom=nom_projet,
        materiaux=mat,
        noeuds=noeuds,
        barres=barres,
        dalles=dalles,
        semelles=semelles,
    )

    return projet, erreurs


# ── Utilitaires ────────────────────────────────────────────────────────────────
def _normaliser_sens(s: str) -> str:
    """Normalise le sens de portée depuis la valeur Excel."""
    s = s.strip().upper()
    if "X" in s and "Y" not in s:
        return "Sens X"
    if "Y" in s and "X" not in s:
        return "Sens Y"
    return "XY"


def valider_coherence(projet: Projet) -> List[str]:
    """
    Vérifications de cohérence supplémentaires après lecture.
    Retourne une liste d'avertissements (non bloquants).
    """
    avertissements = []
    ids_noeuds = {n.id for n in projet.noeuds}
    ids_barres = {b.id for b in projet.barres}

    # Vérifier que les semelles référencent des barres existantes
    for s in projet.semelles:
        if s.id_poteau not in ids_barres:
            avertissements.append(
                f"Semelle : poteau ID={s.id_poteau} inexistant dans les barres"
            )
        if s.long_X_vers > 0 and s.long_X_vers not in ids_barres:
            avertissements.append(
                f"Semelle C{s.id_poteau} : Long_X_vers={s.long_X_vers} "
                "n'existe pas dans les barres"
            )
        if s.long_Y_vers > 0 and s.long_Y_vers not in ids_barres:
            avertissements.append(
                f"Semelle C{s.id_poteau} : Long_Y_vers={s.long_Y_vers} "
                "n'existe pas dans les barres"
            )

    # Vérifier sections minimales BAEL
    for b in projet.barres:
        if b.b < 0.15:
            avertissements.append(
                f"Barre {b.id} : b={b.b*100:.0f}cm < 15cm (min BAEL)"
            )
        if b.h < 0.15:
            avertissements.append(
                f"Barre {b.id} : h={b.h*100:.0f}cm < 15cm (min BAEL)"
            )

    return avertissements
