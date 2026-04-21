"""
export/excel_writer.py
Export des résultats BAEL 91 en fichier Excel
"""
import io
from pathlib import Path
import sys
sys.path.insert(0, str(Path(__file__).parent.parent))

import re as _re_xls

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


def _statut_pot_xls(r, projet):
    """Statut poteau pour export Excel — messages explicites."""
    b = next((b for b in projet.barres if b.id == r.barre_id), None)
    alertes = []
    if r.alerte_am:
        vS = r.vS or ""
        if "As=" in vS and "As_max" in vS:
            if b:
                As_max = 0.05 * b.b * 1000 * b.h * 1000 / 100
                alertes.append(
                    f"As={r.As:.2f}>As_max={As_max:.2f}cm2 augmenter b/h")
        elif "sig=" in vS and ">" in vS:
            m = _re_xls.search(r"sig=([0-9.]+)>([0-9.]+)MPa", vS)
            if m:
                alertes.append(
                    f"sig={m.group(1)}>fbu={m.group(2)}MPa insuff.")
            else:
                alertes.append("Contrainte beton depassee")
        elif "REVOIR" in vS.upper():
            alertes.append(vS)
        else:
            alertes.append("Recouvrement non conforme (>50%)")
    if r.lam > 70:
        alertes.append(f"lam={r.lam:.0f}>70 hors forfaitaire")
    return " | ".join(alertes) if alertes else "OK"



def _statut_pou_xls(r):
    """Statut poutre pour export Excel."""
    import re
    alertes = []
    if r.mu_r > 0.392:
        alertes.append(f"mu={r.mu_r:.3f}>0.392 beton insuff. revoir bxh")
    if r.vCis and "REVOIR" in r.vCis.upper():
        alertes.append(f"Cisaillement excessif")
    if r.vFleche and "REVOIR" in r.vFleche.upper():
        m = re.search(r"f≈([0-9.]+)>([0-9.]+)mm", r.vFleche)
        if m:
            alertes.append(f"Fleche f={m.group(1)}mm>{m.group(2)}mm")
        else:
            alertes.append("Fleche excessive")
    if r.vH and "REVOIR" in r.vH.upper():
        alertes.append(f"h insuff. ({r.vH})")
    return " | ".join(alertes) if alertes else "OK"


def _statut_dal_xls(r):
    """Statut dalle pour export Excel."""
    alertes = []
    if r.alerte:
        if hasattr(r, 'mu_r') and r.mu_r > 0.392:
            alertes.append(f"mu={r.mu_r:.3f}>0.392 section insuff.")
        elif r.vFlex and "REVOIR" in r.vFlex.upper():
            alertes.append(f"Section insuff. ({r.vFlex})")
        if hasattr(r, 'vH') and r.vH and "REVOIR" in r.vH.upper():
            alertes.append(f"h insuff. ({r.vH})")
        if not alertes:
            alertes.append("REVOIR")
    return " | ".join(alertes) if alertes else "OK"


def _statut_sem_xls(s, q_adm):
    """Statut semelle pour export Excel."""
    alertes = []
    if hasattr(s, 'q_min') and s.q_min is not None and s.q_min < 0:
        alertes.append(f"Soulevement q_min={s.q_min:.0f}kN/m2")
    if s.q_max > q_adm * 1.01:
        alertes.append(f"q_max={s.q_max:.0f}>q_adm={q_adm:.0f}kN/m2")
    if s.alerte and not alertes:
        alertes.append(str(s.alerte))
    return " | ".join(alertes) if alertes else "OK"


def exporter_excel(res, projet) -> bytes:
    """Exporte les résultats dans un fichier Excel. Retourne les bytes."""
    if not OPENPYXL_OK:
        raise ImportError("openpyxl non disponible")

    wb = openpyxl.Workbook()

    # Styles
    HDR = Font(bold=True, color="FFFFFF", size=9)
    HDR_FILL = PatternFill("solid", fgColor="1F3864")
    THIN = Side(style="thin", color="CCCCCC")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER = Alignment(horizontal="center", vertical="center")

    def style_hdr(cell, txt, bg="1F3864"):
        cell.value = txt
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.border = BORDER
        cell.alignment = CENTER

    def style_dat(cell, val, bg="FFFFFF"):
        cell.value = val
        cell.font = Font(size=9)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.border = BORDER
        cell.alignment = CENTER

    # ── Feuille Résumé ────────────────────────────────────────────────────────
    ws0 = wb.active
    ws0.title = "Resumé"
    ws0.column_dimensions["A"].width = 30
    ws0.column_dimensions["B"].width = 15

    ws0["A1"] = f"Projet : {projet.nom}"
    ws0["A1"].font = Font(bold=True, size=12, color="1F3864")
    rows_resume = [
        ("Poutres calculées", len(res.poutres)),
        ("Poteaux calculés",  len(res.poteaux)),
        ("Dalles calculées",  len(res.dalles)),
        ("Semelles",          len(res.semelles)),
        ("Alertes poutres",   sum(1 for r in res.poutres if r.alerte)),
        ("Alertes poteaux",   sum(1 for r in res.poteaux if r.alerte_am)),
    ]
    for i, (lab, val) in enumerate(rows_resume, 3):
        ws0[f"A{i}"] = lab
        ws0[f"B{i}"] = val

    # ── Feuille Poutres ───────────────────────────────────────────────────────
    ws1 = wb.create_sheet("Poutres")
    hdrs = ["Niv.", "Élément", "Section", "Mu (kN.m)", "Tu (kN)",
            "As long. (cm²)", "As chap. (cm²)", "As chaîn. (cm²)",
            "At/st (cm²/m)", "Statut"]
    cols_w = [5, 16, 10, 10, 9, 12, 12, 12, 12, 40]
    for c, (h, w) in enumerate(zip(hdrs, cols_w), 1):
        style_hdr(ws1.cell(1, c), h)
        ws1.column_dimensions[get_column_letter(c)].width = w
    ws1.row_dimensions[1].height = 28

    for i, r in enumerate(res.poutres, 2):
        b = next((b for b in projet.barres if b.id == r.barre_id), None)
        niv = b.niveau if b else "?"
        bg = "FCE4D6" if r.alerte else "DEEAF1"
        vals = [niv, r.etiq, r.section, round(r.Mu,2), round(r.Tu,2),
                round(r.As_long,2), round(r.As_chap,2) if r.As_chap else "—",
                round(r.As_chaine,2) if r.As_chaine else "—",
                round(r.At_st,2),
                _statut_pou_xls(r)]
        for c, v in enumerate(vals, 1):
            style_dat(ws1.cell(i, c), v, bg)
        ws1.row_dimensions[i].height = 14

    # ── Feuille Poteaux ───────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Poteaux")
    hdrs2 = ["Niv.", "Élément", "Section", "Nu (kN)", "As (cm²)",
             "α", "λ", "Vérif. ELU", "Statut"]
    cols_w2 = [5, 16, 10, 10, 10, 7, 7, 18, 18, 8]
    for c, (h, w) in enumerate(zip(hdrs2, cols_w2), 1):
        style_hdr(ws2.cell(1, c), h, "375623")
        ws2.column_dimensions[get_column_letter(c)].width = w
    ws2.row_dimensions[1].height = 28

    for i, r in enumerate(res.poteaux, 2):
        b = next((b for b in projet.barres if b.id == r.barre_id), None)
        niv = b.niveau if b else "?"
        bg = "FCE4D6" if r.alerte_am else "E2EFDA"
        vals = [niv, r.etiq, r.section, round(r.Nu,1), round(r.As,2),
                round(r.alpha,2), round(r.lam,0),
                r.vL,
                _statut_pot_xls(r, projet)]
        for c, v in enumerate(vals, 1):
            style_dat(ws2.cell(i, c), v, bg)
        ws2.row_dimensions[i].height = 14

    # ── Feuille Dalles ────────────────────────────────────────────────────────
    ws_d = wb.create_sheet("Dalles")
    hdrs_d = ["ID", "Type", "h/e", "Mu_x (kN.m)", "Mu_y (kN.m)",
              "As nerv. (cm²/m)", "As rép. (cm²/m)", "Statut"]
    for c, h in enumerate(hdrs_d, 1):
        style_hdr(ws_d.cell(1, c), h, "1F3864")
        ws_d.column_dimensions[get_column_letter(c)].width = 14
    ws_d.row_dimensions[1].height = 28
    for i, r in enumerate(res.dalles, 2):
        bg = "FCE4D6" if r.alerte else "EEF4FF"
        vals = [r.dalle_id, r.type_dalle, r.typH,
                round(r.Mu_x, 2),
                round(r.Mu_y, 2) if r.Mu_y > 0 else "—",
                round(r.As_nerv, 2), round(r.As_rep, 2),
                _statut_dal_xls(r)]
        for c, v in enumerate(vals, 1):
            style_dat(ws_d.cell(i, c), v, bg)
        ws_d.row_dimensions[i].height = 14

    # Dictionnaire nom poteau (tous niveaux)
    noms_pots = {b.id: b.nom for b in projet.barres
                 if b.type_elem == "poteau"}
    def _np(bid): return noms_pots.get(bid, f"C{bid}") if bid else "—"

    # ── Feuille Fondations ────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Fondations")
    hdrs3 = ["Poteau", "Type", "B (m)", "L (m)", "e (cm)",
             "Nu (kN)", "q_max (kN/m²)", "Asx (cm²/m)", "Asy (cm²/m)", "Amorces", "Statut"]
    for c, h in enumerate(hdrs3, 1):
        style_hdr(ws3.cell(1, c), h, "7B3F00")
        ws3.column_dimensions[get_column_letter(c)].width = 12
    for i, s in enumerate(res.semelles, 2):
        bg = "FCE4D6" if s.alerte else "FDF0E4"
        vals = [_np(s.id_poteau),
                "Centrée" if s.ex==0 and s.ey==0 else "Excentrique",
                round(s.B,2), round(s.L_sem,2), round(s.e_sem*100,0),
                round(s.Nu_ser,1), round(s.q_max,0),
                round(s.Asx,2), round(s.Asy,2),
                f"{s.nb_amorce}HA{s.phi_amorce} ls={s.ls_amorce*100:.0f}cm",
                _statut_sem_xls(s, projet.materiaux.q_adm)]
        for c, v in enumerate(vals, 1):
            style_dat(ws3.cell(i, c), v, bg)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
